# auto_format_excel     |   tự động định dạng lại file output
# is_empty              |   kiểm tra xem ô dữ liệu có không không, nếu rổng thì trả về true
# normalize_label       |   định dạng dữ liệu thành chữ thường
# check_existing_file   |   kiểm tra tổng quan file input, nếu file không tồn tại sẽ trả về 0, nếu file tồn tại mà rỗng thì trả về 1, 
#                       |   nếu mã đơn hàng đã tồn tại thì trả về 2, File đã tồn tại và sẵn sàng thêm dữ liệu thì trả về 3
# convert_xls_to_xlsx   |   chuyển đổi từ file .xls thành .xlsx
# sheet_one_value       |   tìm dữ liệu trong sheet 1
# safe_float            |   ép kiểu float an toàn, kiểm tra xem dữ liệu hiện tại có ép kiểu được không, nếu được thì mới được phép ép kiểu
# original_cases        |   
# total_cases           |   tính tổng các dữ liệu trong header 'cases book'
# Xu_Ly_File_Input      |   tái định dạng lại file input như là kiểm tra đuôi file, nếu là .xls thì thực hiện chuyển đổi thanh .xlsx, 
#                       |   và kết quả trả về là mã đơn hàng, Cột A và B của sheet 1, mà tên file .xlsx sau khi định dạng
# Xu_Ly_Data_Input      |   trích xuất dữ liệu ở sheet 1 như tên shipper, weight, volume, origin port code và date đã được định dạng về DD/MM/YYYY HH:MM
# Loc_Thong_Tin         |   tiến hành xử lý dữ liệu và thêm dữ liệu cần thiết vào file quản lý hoặc tạo mới nếu chưa có
# Cap_Nhat_Thong_Tin    |   cập nhật dữ liệu mới vào file quản lý
# Gop_File              |   tương tự như loc thong tin nhưng xử lý nhiều file hơn

import os
import re
import pandas as pd
from pathlib import Path
from datetime import datetime
import win32com.client as win32
from openpyxl.styles import Font
from openpyxl import load_workbook

# cần openpyxl để định dạng file excel đầu ra
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Các header cần tìm
shipper_keys = ['shipper']
weight_keys = ['total weight (kg)']
volume_keys = ['total volume (cbm)']
cases_keys = ['cases booked']
opc_keys = ['origin port code']
date_keys = ['date freight available']

# định dạng file output, kết quả trả về là file đã được định dạng
def auto_format_excel(path, sheet_name=None, max_col_width=80, base_width_factor=1.1, line_height=15):
    """
    Mở file với openpyxl và:
    - bật wrap_text cho tất cả ô
    - set column width theo nội dung (giới hạn max_col_width)
    - set row height dựa trên số dòng trong ô (count '\n' + 1) * line_height
    """
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Tính độ rộng tối đa có thể cần cho mỗi cột (dựa vào nội dung)
    col_max_chars = {}
    for row in ws.iter_rows(values_only=True):
        for idx, cell in enumerate(row, start=1):
            if cell is None:
                length = 0
            else:
                # lấy chuỗi, tính max length trong các dòng nếu có newline
                s = str(cell)
                lines = s.splitlines()
                length = max((len(line) for line in lines), default=0)
            col_max_chars[idx] = max(col_max_chars.get(idx, 0), length)

    # Set column widths
    for idx, max_chars in col_max_chars.items():
        col_letter = get_column_letter(idx)
        # Thông thường 1 char ~ 1.0 đơn vị width; nhân thêm hệ số để vừa nhìn
        width = max_chars * base_width_factor + 2
        if width > max_col_width:
            width = max_col_width
        if width < 8:  # tối thiểu
            width = 8
        ws.column_dimensions[col_letter].width = width

    # Wrap text và set alignment, set row height dựa trên số dòng lớn nhất trong hàng
    for r_idx, row in enumerate(ws.iter_rows(), start=1):
        max_lines, i = 1, 0
        for cell in row:
            if i != 1 or r_idx == 1:
                cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(wrapText=True, vertical='top')
            # bật wrap text
            if cell.value is not None:                
                # đếm số dòng
                lines = str(cell.value).count('\n') + 1
                if lines > max_lines:
                    max_lines = lines
            i += 1
        
        # set row height (nhân theo số dòng)
        ws.row_dimensions[r_idx].height = max_lines * line_height

    wb.save(path)
    wb.close()

# kiểm tra ô dữ liệu có rỗng không, kết quả là true hoặc false
def is_empty(v):
    return pd.isna(v) or (isinstance(v, str) and v.strip() == '')

# tái định dạng string về chữ thường, kết quả trả về là rỗng hoặc chuỗi đã được định dạng chữ thường
def normalize_label(s):
    if s is None:
        return ''
    return re.sub(r'[:\s]+$', '', str(s).strip().lower())

# kiểm tra tổng quát file đầu ra, kết quả trả về là mã lỗi 0, 1, 2, 3
def check_existing_file(output_path, ma_don_hang):

    existing_df = ''
    # kiểm tra file có tồn tại hoặc đường dẫn có đúng không
    if os.path.exists(output_path):
        existing_df = pd.read_excel(output_path)
    else:
        return 0
    
    if os.path.getsize(output_path) == 0:
        # File có tồn tại nhưng trống rỗng
        return 1

    # Kiểm tra trùng mã đơn hàng
    for _, row in existing_df.iterrows():
        if ma_don_hang in row.iloc[0]:   # hoặc row["Mã đơn hàng"]
            return 2 
        
    return 3

# hàm chuyển đổi từ file . xls thành .xlsx, không có trả về
def convert_xls_to_xlsx(xls_path, xlsx_path):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(str(xls_path))
    wb.SaveAs(str(xlsx_path), FileFormat=51)  # 51 = .xlsx
    wb.Close()
    excel.Application.Quit()

# tìm kiếm dữ liệu trong sheet 1 (chỉ cột A và B), kết quả trả về là dữ liệu ở header cần tìm
def sheet_one_value(df_ab, key_variants):
    # lấy độ dài của sheet
    n = len(df_ab)
    # duyệt dữ liệu theo chiều dọc của cột A để tìm header
    for i in range(n):
        # gán header cho a
        a = df_ab.iat[i, 0]
        # tái định dạng thành chữ thường cho header để dễ dàng so sánh
        a_norm = normalize_label(a)
        matched = False
        for kv in key_variants:
            # thực hiện so sánh, nếu trúng với header cần tìm thì trả về true và thoát vòng lập
            if a_norm.startswith(kv):
                matched = True
                break
        # khi tìm được header cần tìm thì lấy dữ liệu
        if matched:
            parts = None
            # gán dữ liệu bên cột B cho biến b
            b = df_ab.iat[i, 1]
            # kiểm tra cột B có rỗng không, nếu rỗng thì bỏ qua
            if not is_empty(b):
                parts = str(b).strip()
            else:
                continue
            return parts
    return ''

# kiểm tra ép kiểu float an toàn, nếu ép kiểu thành công thì trả về true, ngược lại thì false
def safe_float(val):
    try:
        float(val)
        return True
    except (TypeError, ValueError):
        return False

# Xử lý trường hợp dữ liệu không phải kiểu số, cụ thể là có từ Original, kết quả trả về là dữ liệu kiểu float
def original_cases(val):
    if not isinstance(val, str):
        return None
    # tách dữ liệu cần thiết ra khỏi từ Original
    match = re.search(r"Original:\s*([\d\.]+)", val)
    if match:
        try:
            # trả dữ liệu về kiểu float thay vì string như ban đầu
            return float(match.group(1))
        except ValueError:
            return None
    return None

# tính tổng cột case book, kết quả trả về là tổng tất cả case book
def total_cases(path, sheet_id, key_variants):
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[sheet_id]
    fv = None
    total = 0
    found = False
    key_col = None
    key_row = None

    # duyệt qua toàn bộ file theo chiều ngang để tìm header 'cases book'
    for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        for c_idx, cell in enumerate(row, start=1):
            val = cell.value
            if is_empty(val):
                continue
            norm = normalize_label(val).replace('\n', ' ')
            # kiểm tra startswith trên từng variant
            for kv in key_variants:
                if norm.startswith(kv):
                    # tìm thấy key
                    found = True
                    key_col = c_idx
                    key_row = r_idx
                    break
            if found:
                break
        if found:
            break

    if not found:
        wb.close()
        return 0  # không thấy key -> trả về 0

    # Duyệt từ hàng tiếp theo (key_row+1) xuống cuối theo cột key_col
    max_row = ws.max_row
    for r in range(key_row + 1, max_row + 1):
        cell = ws.cell(row=r, column=key_col)
        # kiểm tra rỗng
        if is_empty(cell.value):
            continue

        # kiểm tra strikethrough (gạch ngang), nếu dữ liệu có định dàng này thì bỏ qua, không cộng
        strike = False
        try:
            strike = bool(getattr(cell.font, "strike", False))
        except:
            strike = False

        if strike:
            continue

        if safe_float(cell.value):
            # thử chuyển thành float
            fv = float(cell.value)
        else:
            # nếu dữ liệu thuộc string thì xử lý trường hợp Original
            fv = original_cases(cell.value)
        
        # kiểm tra dữ liệu có none hay không trước khi tính tổng
        if fv is not None:
            total += fv
        else:
            # nếu không phải số, bỏ qua
            continue

    wb.close()

    return total

# xữ lý file đầu vào của khách hàng, kết quả trả về là mã đơn hàng, sheet 1 với 2 cột A và B, và tên file cuối cùng sau khi xử lý chuyển đổi
def Xu_Ly_File_Input(input_path):
    p = Path(input_path)
    if not p.exists():
        raise FileNotFoundError(f"File không tồn tại: {input_path}")
    input_path_final = input_path
    read_kwargs = {}

    # kiểm tra đuôi file ban đầu, nếu là .xls thì thực hiện đổi tên file
    if p.suffix.lower() == '.xls':
        # thêm tên file mới với đuôi .xlsx
        input_path_xlsx = p.with_suffix('.xlsx')
        # tạo file mới với tên file mới
        convert_xls_to_xlsx(input_path, input_path_xlsx)
        # gán lại tên file final
        input_path_final = input_path_xlsx
        # xóa file cũ
        os.remove(input_path)

    # Đọc file không lấy header (để A1 là iat[0,0])
    df_1 = pd.read_excel(input_path_final, dtype=object, header=None, sheet_name=0, **read_kwargs)

    # nếu sheet 1 không có đủ 2 cột dữ liệu sẽ báo lỗi
    if df_1.shape[1] < 2:
        raise ValueError("File phải có ít nhất 2 cột (A và B).")

    # copy 2 cột A và B vào biên df_ab
    df_ab = df_1.iloc[:, :2].copy().reset_index(drop=True)

    # Mã đơn hàng lấy ô A1:
    ma_don_hang = ''
    if not is_empty(df_ab.iat[0, 0]):
        ma_don_hang = str(df_ab.iat[0, 0]).strip().split(" ")[0]

    return df_ab, ma_don_hang, input_path_final

# xử lý trích xuất dữ liệu cần thiết từ sheet 1, kết quả trả về là các dữ liệu đó. VD: tên shipper, cân nặng, kích thước, đia điểm và ngày
def Xu_Ly_Data_Input(df_ab):
    # trích xuất dữ liệu từ sheet 1
    shipper_val = sheet_one_value(df_ab, shipper_keys)
    weight_val = sheet_one_value(df_ab, weight_keys)
    volume_val = sheet_one_value(df_ab, volume_keys)
    opc_val = sheet_one_value(df_ab, opc_keys)
    date_val = sheet_one_value(df_ab, date_keys)

    # Nếu date_val là string dạng "09/16/2025 16:30"
    if isinstance(date_val, str) and date_val.strip():
        try:
            dt = datetime.strptime(date_val.strip(), "%m/%d/%Y %H:%M")
            date_val = dt.strftime("%d/%m/%Y %H:%M")  # Xuất lại theo DD/MM/YYYY HH:MM
        except ValueError:
            # Nếu không parse được thì giữ nguyên
            pass
    
    return shipper_val, weight_val, volume_val, opc_val, date_val

# Hàm xử lý chính, lấy dữ liệu và tiến hành thêm vào file quản lý hoặc tạo mới file quản lý nếu chưa có  
def Loc_Thong_Tin(input_path, output_path):
    # df_2 là input sheet 2, df_ab là sheet 1 có 2 cột
    df_ab, ma_don_hang, input_path_xlsx = Xu_Ly_File_Input(input_path)

    # Kiểm tra Mã đơn hàng có trong file output chưa nếu output đã tồn tại
    # Maloi = 0 | File chưa tồn tại
    # Maloi = 1 | File có tồn tại nhưng trống rỗng
    # Maloi = 2 | Trùng mã đơn hàng
    # Maloi = 3 | File đã tồn tại và sẵn sàng thêm dữ liệu
    Maloi = check_existing_file(output_path, ma_don_hang)
    if Maloi == 2:
        return 0
    
    # Các giá trị cần lấy từ file
    shipper_val, weight_val, volume_val, opc_val, date_val = Xu_Ly_Data_Input(df_ab)
    cases_val = total_cases(input_path_xlsx, 1, cases_keys)

    # nếu địa chỉ ở Đà Nẵng thì thêm DAD vào trước mã đơn hàng
    if opc_val == 'DAD':
        ma_don_hang = 'DAD ' + ma_don_hang

    # tạo dữ liệu cho một hàng
    new_row = pd.DataFrame([{
        'Mã đơn hàng': ma_don_hang,
        'Shipper': shipper_val,
        'Origin Port Code': opc_val,
        'Date Freight Available': date_val,
        'Cases Booked' : float(cases_val),
        'Total Weight (KG)': float(weight_val),
        'Total Volume (CBM)': float(volume_val)
        
    }])

    # Nếu file output đã tồn tại thì đọc rồi nối thêm
    if Maloi == 3:
        existing_df = pd.read_excel(output_path)
        out_df = pd.concat([existing_df, new_row], ignore_index=True)
    else:
        out_df = new_row
        
    # lưu file tạm rồi format bằng openpyxl
    out_df.to_excel(output_path, index=False)

    # định dạng tự động
    auto_format_excel(output_path, sheet_name=None, max_col_width=80, base_width_factor=1.1, line_height=18)

    return True

# cập nhật dữ liệu mới so với dữ liệu đã thêm vào trước đó, do có sự thay đổi từ phía khách hàng
def Cap_Nhat_Thong_Tin(input_path, output_path, changed_cells):
    # df_2 là input sheet 2, df_ab là sheet 1 có 2 cột
    df_ab, ma_don_hang, input_path_xlsx = Xu_Ly_File_Input(input_path)
    # Các giá trị cần lấy từ file
    shipper_val, weight_val, volume_val, opc_val, date_val = Xu_Ly_Data_Input(df_ab)
    cases_val = total_cases(input_path_xlsx, 1, cases_keys)
    existing_df = pd.read_excel(output_path)

    if opc_val == 'DAD':
        ma_don_hang = 'DAD ' + ma_don_hang

    # Tìm và cập nhật
    flag = False
    for idx, row in existing_df.iterrows():
        if ma_don_hang == row.iloc[0]:
            # các dữ liệu cần cập nhật
            updates = {
                "Cases Booked": float(cases_val),
                "Total Weight (KG)": float(weight_val),
                "Total Volume (CBM)": float(volume_val)                 
            }
            # tìm kiếm địa chỉ cần cập nhật
            for col, new_val in updates.items():
                # ép kiểu tất cả dữ liệu về float
                if safe_float(row.get(col)):
                    old_val = float(row.get(col))
                
                # tiến hành cập nhật và lưu lại địa chỉ cập nhật để định dạng đánh dấu sau này
                if str(old_val).strip() != str(new_val).strip():
                    existing_df.at[idx, col] = new_val
                    changed_cells.append((idx, col))  # lưu lại ô thay đổi   
            flag = True
            break  # dừng sau khi tìm thấy để tránh ghi đè nhiều dòng
    if not flag:
        return False

    # Lưu lại file
    existing_df.to_excel(output_path, index=False)

    # mở lại bằng openpyxl để định dạng ô đổi màu
    wb = load_workbook(output_path)
    ws = wb.active

    # ánh xạ tên cột -> số cột
    col_map = {col: i+1 for i, col in enumerate(existing_df.columns)}

    for idx, col in changed_cells:
        excel_row = idx + 2  # +2 vì DataFrame index bắt đầu từ 0, Excel từ 1 và có header
        excel_col = col_map[col]
        cell = ws.cell(row=excel_row, column=excel_col)
        cell.font = Font(color="FF0000")  # chữ đỏ

    wb.save(output_path)
    wb.close()

    # định dạng tự động (giữ wrap text, width, height...)
    auto_format_excel(output_path, sheet_name=None, max_col_width=80, base_width_factor=1.1, line_height=18)

    return True

# xử lý lọc dữ liệu với nhiều file hơn vì khách hàng có sự tách file, phải tổng hợp toàn bộ file khách hàng gửi
def Gop_File(list_path_file, output_path):
    # tạo các biến tổng cho toàn bộ file
    ma_don_hang_total = shipper_val = opc_val = date_val = ''
    weight_val_total = volume_val_total = cases_val_total = 0

    # tiến hành duyệt qua danh sách file
    for idx, input_path in enumerate(list_path_file):
        # df_2 là input sheet 2, df_ab là sheet 1 có 2 cột A và B
        df_ab, ma_don_hang, input_path_xlsx = Xu_Ly_File_Input(input_path)
        # Các giá trị cần lấy từ file, tuy nhiên các dữ liệu như tên shipper, địa chỉ gửi và ngày gửi sẽ không có tính tổng bởi vì tất cả các file đều giống nhau
        shipper_val, weight_val, volume_val, opc_val, date_val = Xu_Ly_Data_Input(df_ab)
        cases_val = total_cases(input_path_xlsx, 1, cases_keys)

        # xử lý tên file, nếu là file đầu tiên thì không cần thêm kí tự cộng
        if idx != 0:
            ma_don_hang_total = ma_don_hang_total + ' + ' + ma_don_hang
        else:
            ma_don_hang_total = ma_don_hang
        
        # xử lý tổng các giá trị kiểu số
        cases_val_total += float(cases_val)
        weight_val_total += float(weight_val)
        volume_val_total += float(volume_val)   

    Maloi = check_existing_file(output_path, ma_don_hang_total)
    if Maloi == 2:
        return 0
    
    new_row = pd.DataFrame([{
        'Mã đơn hàng': ma_don_hang_total,
        'Shipper': shipper_val,
        'Origin Port Code': opc_val,
        'Date Freight Available': date_val,
        'Cases Booked' : cases_val_total,
        'Total Weight (KG)': float(weight_val_total),
        'Total Volume (CBM)': float(volume_val_total)
        
    }])

    # Nếu file output đã tồn tại thì đọc rồi nối thêm
    if Maloi == 3:
        existing_df = pd.read_excel(output_path)
        out_df = pd.concat([existing_df, new_row], ignore_index=True)
    else:
        out_df = new_row
        
    # lưu file tạm rồi format bằng openpyxl
    out_df.to_excel(output_path, index=False)

    # định dạng tự động
    auto_format_excel(output_path, sheet_name=None, max_col_width=80, base_width_factor=1.1, line_height=18)

    return True